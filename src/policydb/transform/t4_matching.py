from __future__ import annotations

import re
from datetime import UTC, date, datetime

import polars as pl
from openpyxl.utils import column_index_from_string
from rapidfuzz import fuzz

from policydb.settings import Settings
from policydb.transform.normalization import clean_text, normalize_title, stable_id


def _date(value: object) -> date | None:
    match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", str(value or ""))
    if not match:
        return None
    try:
        return date(*map(int, match.groups()))
    except ValueError:
        return None


def _city(value: object) -> str:
    text = clean_text(value) or ""
    for suffix in ("壮族自治区", "回族自治区", "维吾尔自治区", "自治区", "特别行政区"):
        text = text.replace(suffix, "")
    return text.removesuffix("市").removesuffix("省")


def build_t4_match_candidates(settings: Settings | None = None) -> dict:
    settings = settings or Settings.discover()
    path = next(
        (settings.root / "data" / "staging" / "excel").glob(
            "*T4_2023年城市需求支持政策.parquet"
        )
    )
    staging = pl.read_parquet(path).filter(
        pl.col("source_column").is_in(
            [column_index_from_string(value) for value in ("B", "C", "AB")]
        )
    )
    by_row: dict[int, dict[str, str | None]] = {}
    for cell in staging.iter_rows(named=True):
        by_row.setdefault(cell["source_row"], {})[cell["source_column_letter"]] = cell[
            "cell_value"
        ]
    records = pl.read_parquet(settings.curated / "records.parquet").filter(
        pl.col("source_sheet") == "T1 房地产政策目录"
    )
    linked_features = pl.read_parquet(settings.curated / "policy_features.parquet").filter(
        pl.col("record_id").is_not_null()
    )
    linked_by_row: dict[int, str] = {}
    for row in linked_features.iter_rows(named=True):
        match = re.search(r"\d+", row["source_cell"] or "")
        if match:
            linked_by_row[int(match.group())] = row["record_id"]
    record_rows = records.select(
        "record_id", "title", "title_normalized", "record_date", "geography_original"
    ).iter_rows(named=True)
    record_list = list(record_rows)
    now = datetime.now(UTC).isoformat()
    output: list[dict] = []
    for row_number, cells in sorted(by_row.items()):
        if row_number < 9:
            continue
        city_raw = cells.get("B")
        date_raw = cells.get("C")
        title_raw = cells.get("AB")
        parsed_date = _date(date_raw)
        normalized_title = normalize_title(title_raw)
        linked_id = linked_by_row.get(row_number)
        scored = []
        for record in record_list:
            if _city(city_raw) != _city(record["geography_original"]):
                continue
            day_difference = (
                abs((record["record_date"] - parsed_date).days)
                if record["record_date"] and parsed_date
                else None
            )
            if day_difference is not None and day_difference > 45:
                continue
            title_score = (
                fuzz.WRatio(normalized_title, record["title_normalized"] or "")
                if normalized_title
                else 0
            )
            date_score = max(0.0, 100.0 - (day_difference or 0) * 4) if parsed_date else 0
            score = 0.7 * title_score + 0.3 * date_score if normalized_title else date_score
            if record["record_id"] == linked_id:
                score = 100.0
            if score >= 55:
                scored.append((score, title_score, day_difference, record))
        for score, title_score, day_difference, record in sorted(
            scored, key=lambda item: item[0], reverse=True
        )[:3]:
            exact = record["record_id"] == linked_id
            output.append(
                {
                    "t4_match_id": stable_id(
                        row_number, record["record_id"], prefix="T4MATCH"
                    ),
                    "source_sheet": "T4 2023年城市需求支持政策",
                    "source_row": row_number,
                    "city_raw": city_raw,
                    "publish_date_raw": date_raw,
                    "policy_title_raw": title_raw,
                    "candidate_record_id": record["record_id"],
                    "candidate_title": record["title"],
                    "candidate_date": record["record_date"],
                    "match_method": "existing_exact" if exact else "title_city_date_fuzzy",
                    "match_score": round(score, 2),
                    "title_score": round(title_score, 2),
                    "date_difference_days": day_difference,
                    "review_status": "approved" if exact else "pending",
                    "evidence": (
                        f"城市={city_raw}; 日期差={day_difference}; "
                        f"标题相似度={title_score:.2f}; 候选标题={record['title']}"
                    ),
                    "created_at": now,
                    "updated_at": now,
                }
            )
    schema = {
        "t4_match_id": pl.String,
        "source_sheet": pl.String,
        "source_row": pl.Int64,
        "city_raw": pl.String,
        "publish_date_raw": pl.String,
        "policy_title_raw": pl.String,
        "candidate_record_id": pl.String,
        "candidate_title": pl.String,
        "candidate_date": pl.Date,
        "match_method": pl.String,
        "match_score": pl.Float64,
        "title_score": pl.Float64,
        "date_difference_days": pl.Int64,
        "review_status": pl.String,
        "evidence": pl.String,
        "created_at": pl.String,
        "updated_at": pl.String,
    }
    frame = pl.DataFrame(output, schema=schema).unique(subset=["t4_match_id"])
    frame.write_parquet(settings.curated / "t4_match_candidates.parquet", compression="zstd")
    return {
        "candidate_count": frame.height,
        "approved_existing_count": frame.filter(pl.col("review_status") == "approved").height,
        "pending_candidate_count": frame.filter(pl.col("review_status") == "pending").height,
        "source_row_count": frame["source_row"].n_unique(),
    }
