from __future__ import annotations

import re
import zipfile
from io import BytesIO
from urllib.parse import urljoin

import fitz
import trafilatura
from bs4 import BeautifulSoup
from openpyxl import load_workbook

_END_PUNCTUATION = set("。！？；.!?;：:")
_ATTACHMENT_SUFFIXES = (".pdf", ".doc", ".docx", ".xls", ".xlsx", ".zip", ".rar")
_MAIN_SELECTORS = (
    "article",
    "#UCAP-CONTENT",
    ".TRS_Editor",
    ".article-content",
    ".article_content",
    ".content",
    "main",
)


def _normalise_block(value: str) -> str:
    return re.sub(r"[ \t\u3000]+", " ", value.replace("\r", "")).strip()


def merge_semantic_blocks(blocks: list[dict | str], min_chars: int = 24) -> dict:
    """Join broken DOM/PDF blocks without inventing text.

    A block is joined only when it is short, when the previous block has no sentence-ending
    punctuation, or when a PDF sentence continues on the next page. The returned evidence
    describes every deterministic join.
    """
    merged: list[dict] = []
    repairs: list[dict] = []
    for index, raw in enumerate(blocks):
        item = {"text": raw, "page": None, "kind": "text"} if isinstance(raw, str) else raw
        text = _normalise_block(str(item.get("text") or ""))
        if not text:
            continue
        page = item.get("page")
        if not merged:
            merged.append({"text": text, "page": page, "kind": item.get("kind", "text")})
            continue
        previous = merged[-1]
        cross_page = page is not None and previous.get("page") not in (None, page)
        continues_sentence = previous["text"][-1:] not in _END_PUNCTUATION
        short_fragment = len(text) < min_chars or len(previous["text"]) < min_chars
        heading = item.get("kind") == "heading"
        if not heading and (continues_sentence or (cross_page and short_fragment)):
            separator = "" if cross_page or continues_sentence else "\n"
            previous["text"] += separator + text
            repairs.append(
                {
                    "left_index": index - 1,
                    "right_index": index,
                    "cross_page": cross_page,
                    "reason": "cross_page_sentence" if cross_page else "semantic_continuity",
                }
            )
        elif short_fragment and item.get("kind") not in {"heading", "table"}:
            previous["text"] += "\n" + text
            repairs.append(
                {
                    "left_index": index - 1,
                    "right_index": index,
                    "cross_page": cross_page,
                    "reason": "short_block",
                }
            )
        else:
            merged.append({"text": text, "page": page, "kind": item.get("kind", "text")})
    return {
        "text": "\n\n".join(item["text"] for item in merged),
        "blocks": merged,
        "repairs": repairs,
    }


def _html_parse(body: bytes, base_url: str | None) -> dict:
    html = body.decode("utf-8", errors="replace")
    soup = BeautifulSoup(html, "html.parser")
    for node in soup.select("script,style,noscript,nav,footer,header"):
        node.decompose()
    title = soup.title.get_text(" ", strip=True) if soup.title else None
    candidates = [soup.select_one(selector) for selector in _MAIN_SELECTORS]
    candidates = [node for node in candidates if node is not None]
    root = max(candidates, key=lambda node: len(node.get_text(" ", strip=True))) if candidates else soup
    blocks: list[dict] = []
    for node in root.select("h1,h2,h3,p,li,tr"):
        text = node.get_text(" ", strip=True)
        if text:
            blocks.append(
                {
                    "text": text,
                    "page": 1,
                    "kind": "heading" if node.name.startswith("h") else "table" if node.name == "tr" else "text",
                }
            )
    repaired = merge_semantic_blocks(blocks)
    extracted = repaired["text"] or trafilatura.extract(html) or root.get_text("\n", strip=True)
    attachments = []
    for link in soup.select("a[href]"):
        href = str(link.get("href") or "").strip()
        label = link.get_text(" ", strip=True)
        clean_href = href.lower().split("?", 1)[0]
        if href and (clean_href.endswith(_ATTACHMENT_SUFFIXES) or "附件" in label):
            attachments.append(
                {"url": urljoin(base_url or "", href), "label": label or href, "source": "html_link"}
            )
    script_text = " ".join(node.get_text(" ", strip=True) for node in soup.select("script"))
    dynamic_hint = bool(
        re.search(r"(__NEXT_DATA__|webpack|ajax|iframe|加载中|请开启JavaScript)", html + script_text, re.I)
    )
    completeness = min(1.0, len(extracted) / 800)
    if title and title in extracted:
        completeness = min(1.0, completeness + 0.1)
    status = "parsed" if len(extracted) >= 40 else "partial"
    return {
        "document_type": "html",
        "title": title,
        "full_text": extracted,
        "page_count": None,
        "parse_status": status,
        "text_blocks": repaired["blocks"],
        "repair_actions": repaired["repairs"],
        "attachments": attachments,
        "dynamic_page_hint": dynamic_hint,
        "completeness_score": round(completeness, 4),
    }


def _pdf_parse(body: bytes) -> dict:
    document = fitz.open(stream=BytesIO(body), filetype="pdf")
    blocks: list[dict] = []
    table_count = 0
    for page_number, page in enumerate(document, start=1):
        for block in page.get_text("blocks"):
            text = _normalise_block(str(block[4]))
            if text:
                blocks.append({"text": text, "page": page_number, "kind": "text"})
        try:
            tables = page.find_tables().tables
        except (AttributeError, RuntimeError, ValueError):
            tables = []
        for table in tables:
            rows = table.extract()
            table_text = "\n".join(
                " | ".join(str(cell or "").strip() for cell in row) for row in rows
            ).strip()
            if table_text:
                blocks.append({"text": table_text, "page": page_number, "kind": "table"})
                table_count += 1
    repaired = merge_semantic_blocks(blocks)
    embedded = [
        {"url": None, "label": name, "source": "pdf_embedded"}
        for name in document.embfile_names()
    ]
    text = repaired["text"]
    completeness = min(1.0, len(text) / max(500, len(document) * 300))
    return {
        "document_type": "pdf",
        "title": document.metadata.get("title") or None,
        "full_text": text,
        "page_count": len(document),
        "parse_status": "parsed" if len(text) >= 40 else "partial",
        "text_blocks": repaired["blocks"],
        "repair_actions": repaired["repairs"],
        "attachments": embedded,
        "dynamic_page_hint": False,
        "table_count": table_count,
        "completeness_score": round(completeness, 4),
    }


def extract_pdf_embedded(body: bytes) -> list[tuple[str, bytes]]:
    """Return embedded files verbatim so the Raw layer can store them append-only."""
    document = fitz.open(stream=BytesIO(body), filetype="pdf")
    return [(name, document.embfile_get(name)) for name in document.embfile_names()]


def _office_parse(body: bytes, content_type: str) -> dict:
    blocks: list[dict] = []
    if "spreadsheet" in content_type or "excel" in content_type:
        workbook = load_workbook(BytesIO(body), read_only=True, data_only=False)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    blocks.append({"text": " | ".join(values), "page": None, "kind": "table"})
        document_type = "xlsx"
    else:
        with zipfile.ZipFile(BytesIO(body)) as archive:
            xml = archive.read("word/document.xml")
        soup = BeautifulSoup(xml, "xml")
        for paragraph in soup.find_all("w:p"):
            text = "".join(node.get_text() for node in paragraph.find_all("w:t")).strip()
            if text:
                blocks.append({"text": text, "page": None, "kind": "text"})
        document_type = "docx"
    repaired = merge_semantic_blocks(blocks)
    return {
        "document_type": document_type,
        "title": None,
        "full_text": repaired["text"],
        "page_count": None,
        "parse_status": "parsed" if repaired["text"] else "partial",
        "text_blocks": repaired["blocks"],
        "repair_actions": repaired["repairs"],
        "attachments": [],
        "dynamic_page_hint": False,
        "completeness_score": min(1.0, len(repaired["text"]) / 500),
    }


def parse_document(body: bytes, content_type: str | None, base_url: str | None = None) -> dict:
    content_type = (content_type or "").lower()
    try:
        if "pdf" in content_type or body.startswith(b"%PDF"):
            return _pdf_parse(body)
        if any(token in content_type for token in ("wordprocessingml", "spreadsheetml", "ms-excel")):
            return _office_parse(body, content_type)
        return _html_parse(body, base_url)
    except (RuntimeError, ValueError, TypeError) as error:
        return {
            "document_type": "pdf" if body.startswith(b"%PDF") else "html",
            "title": None,
            "full_text": "",
            "page_count": None,
            "parse_status": "parser_error",
            "text_blocks": [],
            "repair_actions": [],
            "attachments": [],
            "dynamic_page_hint": False,
            "completeness_score": 0.0,
            "parser_error": f"{type(error).__name__}: {error}",
        }
