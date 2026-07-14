from __future__ import annotations

import csv
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

from policydb.settings import Settings

TARGET_SHEETS = {
    "T1 房地产政策目录",
    "T4 2023年城市需求支持政策",
    "T5 供给侧措施",
    "T6 城市限售政策汇总",
    "T7 中央经济工作会议",
    "T8 中央政治局会议",
    "T9 全国住建工作会议",
    "T10 政府工作报告",
    "房地产项目白名单（城市情况）",
    "房地产项目白名单（企业情况）",
    "PSL专项贷款",
}


def export_excel_compatible(
    template: Path,
    output: Path,
    settings: Settings | None = None,
) -> dict:
    settings = settings or Settings.discover()
    raw_root = (settings.root / "data" / "raw").resolve()
    output = output.resolve()
    if output == template.resolve() or raw_root in output.parents:
        raise ValueError("Excel export must not overwrite or write into Raw")
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(template, data_only=False)
    records = pl.read_parquet(settings.curated / "records.parquet")
    mapping_rows: list[dict] = []

    t1 = workbook["T1 房地产政策目录"]
    if t1.max_column < 9:
        t1.cell(1, 9, "policy_id")
    else:
        t1.cell(1, 9, "policy_id")
    t1_records = records.filter(pl.col("source_sheet") == "T1 房地产政策目录").sort(
        "source_row"
    )
    for record in t1_records.iter_rows(named=True):
        row = int(record["source_row"])
        values = [
            record["record_date"],
            record["geography_original"],
            record["title"],
            record["legacy_category"],
            record["summary"],
            record["full_text"],
            record["primary_source_url"],
            record["notes"],
            record["record_id"],
        ]
        for column, value in enumerate(values, start=1):
            t1.cell(row, column, value)
        mapping_rows.append(
            {
                "source_sheet": "T1 房地产政策目录",
                "source_row": row,
                "export_row": row,
                "policy_id": record["record_id"],
            }
        )

    by_sheet_row = {
        (row["source_sheet"], int(row["source_row"])): row["record_id"]
        for row in records.select("record_id", "source_sheet", "source_row").iter_rows(
            named=True
        )
        if row["source_sheet"] and row["source_row"]
    }
    for sheet_name in TARGET_SHEETS - {"T1 房地产政策目录"}:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        id_column = sheet.max_column + 1
        sheet.cell(1, id_column, "policy_id")
        for row in range(2, sheet.max_row + 1):
            policy_id = by_sheet_row.get((sheet_name, row))
            if policy_id:
                sheet.cell(row, id_column, policy_id)
                mapping_rows.append(
                    {
                        "source_sheet": sheet_name,
                        "source_row": row,
                        "export_row": row,
                        "policy_id": policy_id,
                    }
                )
    if "导出说明" in workbook.sheetnames:
        del workbook["导出说明"]
    note = workbook.create_sheet("导出说明", 0)
    note.append(["项目", "说明"])
    note.append(["数据来源", "Curated政策实体与只读原始Excel模板"])
    note.append(["Raw保护", "本文件是Release导出，不覆盖原始工作簿"])
    note.append(["policy_id", "用于连接旧版工作表行号与规范政策实体"])
    note.append(["派生字段", "研究统计应以Curated/Research层视图为准，不作为人工输入"])
    workbook.save(output)
    mapping_path = output.with_name(output.stem + "_row_mapping.csv")
    with mapping_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(mapping_rows[0]))
        writer.writeheader()
        writer.writerows(mapping_rows)
    return {
        "output": str(output),
        "row_mapping": str(mapping_path),
        "mapped_policy_rows": len(mapping_rows),
        "template_unchanged": True,
    }
