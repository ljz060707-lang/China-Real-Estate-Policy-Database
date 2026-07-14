from __future__ import annotations

import hashlib
import json
import re
import shutil
from datetime import UTC, date, datetime
from pathlib import Path

import polars as pl
import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from policydb.classify.rules import classify, infer_direction
from policydb.settings import Settings
from policydb.transform.normalization import (
    clean_text,
    content_hash,
    normalize_title,
    normalize_url,
    stable_id,
)
from policydb.transform.t4_matching import build_t4_match_candidates

RECORD_COLUMNS = [
    "record_id",
    "record_type",
    "title",
    "title_normalized",
    "record_date",
    "publication_date",
    "issuance_date",
    "effective_date",
    "expiry_date",
    "record_date_original",
    "status",
    "direction",
    "summary",
    "full_text",
    "language",
    "official_level",
    "official_status",
    "source_quality",
    "primary_source_url",
    "landing_page_url",
    "document_url",
    "retrieved_at",
    "content_hash",
    "source_file",
    "source_sheet",
    "source_row",
    "import_batch_id",
    "created_at",
    "updated_at",
    "manual_review_status",
    "notes",
    "geography_original",
    "legacy_category",
]


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def slug_sheet(name: str) -> str:
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff._-]+", "_", name).strip("_")[:100]


def parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value:
        match = re.search(r"(20\d{2}|19\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", str(value))
        if match:
            try:
                return date(*map(int, match.groups()))
            except ValueError:
                pass
    return None


def source_status(url: str | None, text: str) -> tuple[str, int]:
    combined = f"{url or ''} {text}"
    if any(x in combined for x in ("内部", "传闻")):
        return ("internal_unverified" if "内部" in combined else "rumour", 0)
    if url and any(
        x in url for x in ("gov.cn", "pbc.gov.cn", "mohurd.gov.cn", "nfra.gov.cn", "gjj.")
    ):
        return "official", 5
    if url and any(x in url for x in ("xinhuanet", "people.com.cn", "cnstock", "cls.cn")):
        return "authoritative_media", 3
    return ("general_media", 2) if url else ("unknown", 0)


def inventory_excel(path: Path) -> dict:
    wb = load_workbook(path, data_only=False, read_only=False)
    sheets = []
    for idx, ws in enumerate(wb.worksheets, 1):
        rows, cols, formulas = set(), set(), 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    rows.add(cell.row)
                    cols.add(cell.column)
                    formulas += int(cell.data_type == "f")
        sheets.append(
            {
                "sheet_index": idx,
                "sheet_name": ws.title,
                "state": ws.sheet_state,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "nonempty_row_count": len(rows),
                "nonempty_column_count": len(cols),
                "formula_count": formulas,
                "merged_range_count": len(ws.merged_cells.ranges),
                "merged_ranges": [str(x) for x in ws.merged_cells.ranges],
            }
        )
    return {
        "source": str(path),
        "sha256": file_hash(path),
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def _write_staging(wb, source: Path, sha: str, batch: str, out: Path) -> dict[str, int]:
    out.mkdir(parents=True, exist_ok=True)
    counts = {}
    for idx, ws in enumerate(wb.worksheets, 1):
        merged_lookup = {}
        for merged in ws.merged_cells.ranges:
            anchor = ws.cell(merged.min_row, merged.min_col).coordinate
            for row in range(merged.min_row, merged.max_row + 1):
                for col in range(merged.min_col, merged.max_col + 1):
                    merged_lookup[f"{get_column_letter(col)}{row}"] = (str(merged), anchor)
        rows = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None and cell.coordinate not in merged_lookup:
                    continue
                merged_range, merged_anchor = merged_lookup.get(cell.coordinate, (None, None))
                rows.append(
                    {
                        "source_sheet_name": ws.title,
                        "sheet_index": idx,
                        "sheet_state": ws.sheet_state,
                        "source_row": cell.row,
                        "source_column": cell.column,
                        "source_column_letter": get_column_letter(cell.column),
                        "source_cell": cell.coordinate,
                        "original_field_name": clean_text(ws.cell(1, cell.column).value),
                        "cell_value": clean_text(cell.value),
                        "cell_data_type": cell.data_type,
                        "is_formula": cell.data_type == "f",
                        "formula_text": str(cell.value) if cell.data_type == "f" else None,
                        "is_merged": merged_range is not None,
                        "merged_range": merged_range,
                        "merged_anchor": merged_anchor,
                        "import_batch_id": batch,
                        "source_file": source.name,
                        "source_file_sha256": sha,
                    }
                )
        frame = (
            pl.DataFrame(rows, infer_schema_length=None)
            if rows
            else pl.DataFrame({"source_sheet_name": pl.Series([], dtype=pl.String)})
        )
        frame.write_parquet(out / f"{idx:02d}_{slug_sheet(ws.title)}.parquet", compression="zstd")
        counts[ws.title] = len(rows)
    return counts


def _main_records(
    wb, source: Path, sha: str, batch: str, now: datetime
) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    ws = wb["T1 房地产政策目录"]
    records, geos, terms, legacy = [], [], [], []
    aliases = yaml.safe_load(
        (Settings.discover().root / "config" / "geography_aliases.yml").read_text(encoding="utf-8")
    )["aliases"]
    for row in range(2, ws.max_row + 1):
        vals = [ws.cell(row, col).value for col in range(1, 9)]
        if not any(v is not None for v in vals):
            continue
        raw_date, geography, title, category, summary, full_text, url, notes = vals
        parsed = parse_date(raw_date)
        normalized_url = normalize_url(url)
        text = " ".join(clean_text(x) or "" for x in (title, category, summary, full_text))
        rid = stable_id(parsed, geography, title, normalized_url, prefix="POL")
        off_status, quality = source_status(normalized_url, text)
        if "征求意见" in text:
            off_status, status = "consultation_draft", "consultation"
        else:
            status = "issued"
        records.append(
            dict(
                zip(
                    RECORD_COLUMNS,
                    [
                        rid,
                        "policy_document",
                        clean_text(title),
                        normalize_title(title),
                        parsed,
                        None,
                        parsed,
                        None,
                        None,
                        clean_text(raw_date),
                        status,
                        infer_direction(text),
                        clean_text(summary),
                        clean_text(full_text),
                        "zh-CN",
                        "local" if geography and geography != "全国" else "central",
                        off_status,
                        quality,
                        normalized_url,
                        None,
                        None,
                        now,
                        content_hash(title, summary, full_text),
                        source.name,
                        ws.title,
                        row,
                        batch,
                        now,
                        now,
                        "pending" if not title or not normalized_url else "unreviewed",
                        clean_text(notes),
                        clean_text(geography),
                        clean_text(category),
                    ],
                    strict=True,
                )
            )
        )
        original = clean_text(geography)
        standardized = aliases.get(original, original)
        geos.append(
            {
                "record_id": rid,
                "jurisdiction_id": stable_id(standardized, prefix="JUR"),
                "geography_original": original,
                "jurisdiction_name": standardized,
                "relation_type": "issuing_jurisdiction",
                "match_method": "alias" if standardized != original else "exact",
                "match_confidence": 1.0 if standardized else 0.0,
            }
        )
        for item in classify(text):
            tid = stable_id("topic", item["topic"], prefix="TERM")
            terms.append(
                {
                    "record_id": rid,
                    "term_id": tid,
                    "taxonomy_name": "topic",
                    "term_name": item["topic"],
                    "classification_source": "rule",
                    "confidence": item["confidence"],
                    "evidence_excerpt": item["evidence_excerpt"],
                    "review_status": "unreviewed",
                }
            )
        if category:
            legacy.append(
                {
                    "record_id": rid,
                    "annotation_type": "legacy_category",
                    "annotation_value": clean_text(category),
                    "source_sheet": ws.title,
                    "source_row": row,
                }
            )
    return records, geos, terms, legacy


def _topic_events(wb, source: Path, batch: str, now: datetime) -> list[dict]:
    specs = {
        "T5 供给侧措施": "policy_document",
        "T7 中央经济工作会议": "meeting_statement",
        "T8 中央政治局会议": "meeting_statement",
        "T9 全国住建工作会议": "meeting_statement",
        "T10 政府工作报告": "government_report",
        "T11 2007年以来历届全国党代会": "meeting_statement",
        "T12 央行、银保监会、证监会、住建部2014年至今政策梳理": "policy_document",
        "房地产项目白名单（城市情况）": "programme_event",
        "房地产项目白名单（企业情况）": "enterprise_event",
        "PSL专项贷款": "financing_event",
        "附录1 疫情期间政策跟踪": "policy_document",
    }
    output = []
    for sheet_name, kind in specs.items():
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            values = [ws.cell(row, col).value for col in range(1, min(ws.max_column, 12) + 1)]
            if not any(v is not None for v in values):
                continue
            strings = [clean_text(v) for v in values]
            raw_date = next((v for v in values if parse_date(v)), None)
            parsed = parse_date(raw_date)
            url = next(
                (
                    normalize_url(v)
                    for v in values
                    if isinstance(v, str) and v.startswith(("http://", "https://"))
                ),
                None,
            )
            text = " | ".join(v for v in strings if v)
            title = next(
                (v for v in strings if v and not v.startswith("http") and len(v) > 5),
                f"{sheet_name}第{row}行",
            )
            rid = stable_id(sheet_name, row, parsed, title, prefix="REC")
            off_status, quality = source_status(url, text)
            output.append(
                dict(
                    zip(
                        RECORD_COLUMNS,
                        [
                            rid,
                            kind,
                            title[:500],
                            normalize_title(title[:500]),
                            parsed,
                            None,
                            parsed,
                            None,
                            None,
                            clean_text(raw_date),
                            "historical" if parsed and parsed.year < 2020 else "issued",
                            infer_direction(text),
                            text[:2000],
                            text,
                            "zh-CN",
                            "central"
                            if any(x in sheet_name for x in ("中央", "全国", "央行"))
                            else "local",
                            off_status,
                            quality,
                            url,
                            None,
                            None,
                            now,
                            content_hash(text),
                            source.name,
                            sheet_name,
                            row,
                            batch,
                            now,
                            now,
                            "unreviewed",
                            None,
                            None,
                            None,
                        ],
                        strict=True,
                    )
                )
            )
    return output


def import_excel(path: str | Path, settings: Settings | None = None) -> dict:
    settings = settings or Settings.discover()
    source = Path(path).resolve()
    if not source.exists():
        raise FileNotFoundError(source)
    sha = file_hash(source)
    batch = f"excel_{datetime.now(UTC):%Y%m%dT%H%M%SZ}_{sha[:8]}"
    now = datetime.now(UTC)
    seed_dir = settings.root / "data" / "raw" / "seed"
    seed_dir.mkdir(parents=True, exist_ok=True)
    raw_copy = seed_dir / source.name
    if source != raw_copy.resolve():
        if raw_copy.exists() and file_hash(raw_copy) != sha:
            raw_copy = seed_dir / f"{source.stem}_{sha[:8]}{source.suffix}"
        if not raw_copy.exists():
            shutil.copy2(source, raw_copy)
    if file_hash(raw_copy) != sha:
        raise RuntimeError("Raw copy hash mismatch")
    wb = load_workbook(raw_copy, data_only=False, read_only=False)
    inventory = inventory_excel(raw_copy)
    staging_counts = _write_staging(
        wb, raw_copy, sha, batch, settings.root / "data" / "staging" / "excel"
    )
    records, geos, terms, legacy = _main_records(wb, raw_copy, sha, batch, now)
    records.extend(_topic_events(wb, raw_copy, batch, now))
    settings.curated.mkdir(parents=True, exist_ok=True)
    pl.DataFrame(records, infer_schema_length=None).write_parquet(
        settings.curated / "records.parquet"
    )
    pl.DataFrame(geos).write_parquet(settings.curated / "record_jurisdictions.parquet")
    pl.DataFrame(terms).write_parquet(settings.curated / "record_terms.parquet")
    pl.DataFrame(legacy).write_parquet(settings.curated / "legacy_annotations.parquet")
    _write_reference_tables(settings, geos, terms, records, wb, batch, now)
    from policydb.transform.collections import build_collection_layer

    collection_report = build_collection_layer(settings)
    t4_match_report = build_t4_match_candidates(settings)
    manifest = {
        "import_batch_id": batch,
        "source_file": raw_copy.name,
        "source_sha256": sha,
        "imported_at": now.isoformat(),
        "sheet_count": len(wb.sheetnames),
        "main_policy_count": sum(r["source_sheet"] == "T1 房地产政策目录" for r in records),
        "record_count": len(records),
        "staging_cell_counts": staging_counts,
        "inventory": inventory,
        "collection_coverage": collection_report,
        "t4_matching": t4_match_report,
    }
    (settings.root / "data" / "staging" / "import_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    from policydb.query.database import build_database

    build_database(settings)
    return manifest


def _empty_or(rows: list[dict], schema: dict[str, pl.DataType], path: Path) -> None:
    (
        pl.DataFrame(rows, infer_schema_length=None) if rows else pl.DataFrame(schema=schema)
    ).write_parquet(path)


def _write_reference_tables(
    settings: Settings,
    geos: list[dict],
    terms: list[dict],
    records: list[dict],
    wb,
    batch: str,
    now: datetime,
) -> None:
    unique_geo = {}
    unique_terms = {}
    for g in geos:
        unique_geo[g["jurisdiction_id"]] = {
            "jurisdiction_id": g["jurisdiction_id"],
            "name": g["jurisdiction_name"],
            "name_full": g["jurisdiction_name"],
            "name_normalized": g["jurisdiction_name"],
            "administrative_code": None,
            "level": "unknown",
            "parent_id": None,
            "province": None,
            "prefecture": g["jurisdiction_name"],
            "county": None,
            "longitude": None,
            "latitude": None,
            "valid_from": None,
            "valid_to": None,
        }
    for t in terms:
        unique_terms[t["term_id"]] = {
            "term_id": t["term_id"],
            "taxonomy_name": t["taxonomy_name"],
            "term_code": t["term_id"],
            "term_name": t["term_name"],
            "parent_term_id": None,
            "description": None,
            "valid_from": None,
            "valid_to": None,
        }
    _empty_or(list(unique_geo.values()), {}, settings.curated / "jurisdictions.parquet")
    _empty_or(list(unique_terms.values()), {}, settings.curated / "taxonomy_terms.parquet")
    policies = [
        {
            "record_id": r["record_id"],
            "document_number": None,
            "policy_type": r["legacy_category"],
            "policy_level": r["official_level"],
            "mandatory_strength": 1 if r["record_type"] == "policy_document" else 0,
            "implementation_actor": None,
            "target_group": None,
            "policy_goal": None,
            "implementation_method": None,
            "legal_basis": None,
            "quantitative_target_text": None,
            "validity_text": None,
        }
        for r in records
        if r["record_type"] == "policy_document"
    ]
    _empty_or(policies, {}, settings.curated / "policies.parquet")
    documents = [
        {
            "document_id": stable_id(r["record_id"], r["primary_source_url"], prefix="DOC"),
            "record_id": r["record_id"],
            "document_type": "webpage",
            "original_filename": None,
            "local_path": None,
            "source_url": r["primary_source_url"],
            "mime_type": "text/html",
            "sha256": None,
            "retrieved_at": r["retrieved_at"],
            "http_status": None,
            "etag": None,
            "last_modified": None,
            "text_extraction_status": "from_excel",
            "ocr_status": "not_applicable",
        }
        for r in records
        if r["primary_source_url"]
    ]
    versions = [
        {
            "version_id": stable_id(r["record_id"], r["content_hash"], prefix="VER"),
            "record_id": r["record_id"],
            "version_number": 1,
            "valid_from": r["record_date"],
            "valid_to": None,
            "content_hash": r["content_hash"],
        }
        for r in records
    ]
    programme = [
        {
            "record_id": r["record_id"],
            "programme_type": "psl" if r["source_sheet"] == "PSL专项贷款" else "white_list",
            "project_count": None,
            "enterprise_count": None,
            "financing_demand": None,
            "approved_amount": None,
            "disbursed_amount": None,
            "housing_units": None,
            "floor_area": None,
            "project_names": None,
        }
        for r in records
        if r["source_sheet"] == "PSL专项贷款" or r["source_sheet"].startswith("房地产项目白名单")
    ]
    city_rules = _extract_city_policy_rules(wb)
    tier_attributes = _extract_city_tiers(wb)
    demand_features, demand_measures = _extract_demand_features(wb, records)
    update_rows = [
        {
            "run_id": batch,
            "run_type": "excel_import",
            "started_at": now,
            "finished_at": now,
            "source_count": 1,
            "new_record_count": len(records),
            "updated_record_count": 0,
            "unchanged_record_count": 0,
            "failed_record_count": 0,
            "review_required_count": sum(r["manual_review_status"] == "pending" for r in records),
            "code_version": "0.1.0",
            "data_version": settings.data_version,
            "log_path": "data/staging/import_manifest.json",
        }
    ]
    empty_defs = {
        "documents": {
            "document_id": pl.String,
            "record_id": pl.String,
            "document_type": pl.String,
            "original_filename": pl.String,
            "local_path": pl.String,
            "source_url": pl.String,
            "mime_type": pl.String,
            "sha256": pl.String,
            "retrieved_at": pl.Datetime,
            "http_status": pl.Int32,
            "etag": pl.String,
            "last_modified": pl.String,
            "text_extraction_status": pl.String,
            "ocr_status": pl.String,
        },
        "organizations": {
            "organization_id": pl.String,
            "name_original": pl.String,
            "name_standardized": pl.String,
            "organization_type": pl.String,
            "administrative_level": pl.String,
            "parent_organization_id": pl.String,
            "valid_from": pl.Date,
            "valid_to": pl.Date,
        },
        "record_organizations": {
            "record_id": pl.String,
            "organization_id": pl.String,
            "role": pl.String,
        },
        "policy_relations": {
            "relation_id": pl.String,
            "source_record_id": pl.String,
            "target_record_id": pl.String,
            "relation_type": pl.String,
            "evidence": pl.String,
            "confidence": pl.Float64,
        },
        "quantitative_measures": {
            "measure_id": pl.String,
            "record_id": pl.String,
            "measure_type": pl.String,
            "value": pl.Float64,
            "value_min": pl.Float64,
            "value_max": pl.Float64,
            "unit": pl.String,
            "currency": pl.String,
            "population_scope": pl.String,
            "housing_scope": pl.String,
            "condition_text": pl.String,
            "period_start": pl.Date,
            "period_end": pl.Date,
            "evidence_excerpt": pl.String,
        },
        "programme_events": {
            "record_id": pl.String,
            "programme_type": pl.String,
            "project_count": pl.Int64,
            "enterprise_count": pl.Int64,
            "financing_demand": pl.Float64,
            "approved_amount": pl.Float64,
            "disbursed_amount": pl.Float64,
            "housing_units": pl.Int64,
            "floor_area": pl.Float64,
            "project_names": pl.String,
        },
        "policy_versions": {
            "version_id": pl.String,
            "record_id": pl.String,
            "version_number": pl.Int32,
            "valid_from": pl.Date,
            "valid_to": pl.Date,
            "content_hash": pl.String,
        },
        "policy_provisions": {
            "provision_id": pl.String,
            "record_id": pl.String,
            "provision_type": pl.String,
            "provision_text": pl.String,
        },
        "policy_features": {
            "record_id": pl.String,
            "feature_name": pl.String,
            "feature_value": pl.String,
            "source_sheet": pl.String,
            "source_cell": pl.String,
        },
        "jurisdiction_attributes": {
            "jurisdiction_id": pl.String,
            "attribute_name": pl.String,
            "attribute_value": pl.String,
            "valid_from": pl.Date,
            "source": pl.String,
        },
        "city_policy_rules": {
            "jurisdiction_id": pl.String,
            "policy_dimension": pl.String,
            "population_group": pl.String,
            "housing_count": pl.String,
            "loan_status": pl.String,
            "rule_text": pl.String,
            "effective_date": pl.Date,
            "source_cell": pl.String,
            "source_sheet": pl.String,
        },
        "policy_change_comparisons": {
            "record_id": pl.String,
            "jurisdiction_id": pl.String,
            "policy_topic": pl.String,
            "before_text": pl.String,
            "after_text": pl.String,
            "change_summary": pl.String,
            "direction": pl.String,
            "effective_date": pl.Date,
            "source": pl.String,
        },
        "update_runs": {
            "run_id": pl.String,
            "run_type": pl.String,
            "started_at": pl.Datetime,
            "finished_at": pl.Datetime,
            "source_count": pl.Int64,
            "new_record_count": pl.Int64,
            "updated_record_count": pl.Int64,
            "unchanged_record_count": pl.Int64,
            "failed_record_count": pl.Int64,
            "review_required_count": pl.Int64,
            "code_version": pl.String,
            "data_version": pl.String,
            "log_path": pl.String,
        },
    }
    populated = {
        "documents": documents,
        "programme_events": programme,
        "policy_versions": versions,
        "policy_features": demand_features,
        "quantitative_measures": demand_measures,
        "jurisdiction_attributes": tier_attributes,
        "city_policy_rules": city_rules,
        "update_runs": update_rows,
    }
    for name, schema in empty_defs.items():
        _empty_or(populated.get(name, []), schema, settings.curated / f"{name}.parquet")


def _extract_city_policy_rules(wb) -> list[dict]:
    ws = wb["T2 城市房地产政策现状"]
    city_by_column: dict[int, str | None] = {}
    current_city = None
    for column in range(1, ws.max_column + 1):
        heading = clean_text(ws.cell(1, column).value)
        if heading and "政策现状" in heading:
            current_city = heading.split("政策现状", 1)[0].replace("（", "").strip()
        city_by_column[column] = current_city
    output = []
    for row in range(3, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            value = clean_text(ws.cell(row, column).value)
            city = city_by_column.get(column)
            if not value or not city or column <= 2:
                continue
            dimension = clean_text(ws.cell(2, column).value) or clean_text(ws.cell(3, column).value)
            output.append(
                {
                    "jurisdiction_id": stable_id(city, prefix="JUR"),
                    "policy_dimension": dimension,
                    "population_group": None,
                    "housing_count": None,
                    "loan_status": None,
                    "rule_text": value,
                    "effective_date": None,
                    "source_cell": ws.cell(row, column).coordinate,
                    "source_sheet": ws.title,
                }
            )
    return output


def _extract_city_tiers(wb) -> list[dict]:
    ws = wb["能级划分"]
    output = []
    for row in range(1, ws.max_row + 1):
        city, tier = clean_text(ws.cell(row, 1).value), clean_text(ws.cell(row, 2).value)
        if not city or not tier or "城市" in city:
            continue
        output.append(
            {
                "jurisdiction_id": stable_id(city, prefix="JUR"),
                "attribute_name": "city_tier",
                "attribute_value": tier,
                "valid_from": None,
                "source": f"{ws.title}!A{row}:B{row}",
            }
        )
    return output


def _extract_demand_features(wb, records: list[dict]) -> tuple[list[dict], list[dict]]:
    ws = wb["T4 2023年城市需求支持政策"]
    index: dict[tuple[str | None, date | None], list[str]] = {}
    for record in records:
        if record["source_sheet"] != "T1 房地产政策目录":
            continue
        index.setdefault((record["geography_original"], record["record_date"]), []).append(
            record["record_id"]
        )
    columns = {
        4: "purchase_restriction",
        5: "sale_restriction",
        6: "commercial_mortgage_downpayment",
        7: "provident_fund_downpayment",
        8: "provident_fund_limit",
        9: "provident_fund_other_conditions",
        10: "talent_settlement",
        11: "purchase_subsidy_other",
        19: "with_official_document",
        20: "without_official_document",
        22: "county_only_relaxation",
        23: "city_tier",
        28: "official_source_note",
    }
    features, measures = [], []
    for row in range(9, 2119):
        city = clean_text(ws.cell(row, 2).value)
        event_date = parse_date(ws.cell(row, 3).value)
        candidates = index.get((city, event_date), [])
        linked_id = candidates[0] if len(candidates) == 1 else None
        for column, feature_name in columns.items():
            cell = ws.cell(row, column)
            if cell.data_type == "f":
                continue
            value = clean_text(cell.value)
            if value is None:
                continue
            features.append(
                {
                    "record_id": linked_id,
                    "feature_name": feature_name,
                    "feature_value": value,
                    "source_sheet": ws.title,
                    "source_cell": cell.coordinate,
                }
            )
            if linked_id and column in (6, 7, 8):
                match = re.search(r"(?<!\d)(\d+(?:\.\d+)?)", value.replace(",", ""))
                if match:
                    unit = "%" if "%" in value else ("万元" if "万" in value else None)
                    measures.append(
                        {
                            "measure_id": stable_id(linked_id, cell.coordinate, prefix="MEA"),
                            "record_id": linked_id,
                            "measure_type": feature_name,
                            "value": float(match.group(1)),
                            "value_min": None,
                            "value_max": None,
                            "unit": unit,
                            "currency": "CNY" if unit == "万元" else None,
                            "population_scope": None,
                            "housing_scope": None,
                            "condition_text": value,
                            "period_start": None,
                            "period_end": None,
                            "evidence_excerpt": value,
                        }
                    )
    return features, measures
