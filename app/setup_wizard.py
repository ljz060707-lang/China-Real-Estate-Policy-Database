from __future__ import annotations

import hashlib
import os
from io import BytesIO
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook


def initial_setup_status(root: Path) -> dict[str, bool]:
    curated = root / "data" / "curated"
    return {
        "database_ready": (root / "database" / "policydb.duckdb").is_file(),
        "curated_ready": curated.is_dir() and (curated / "records.parquet").is_file(),
    }


def needs_initial_setup(root: Path) -> bool:
    return not all(initial_setup_status(root).values())


def _workbook_summary(content: bytes) -> tuple[int, list[str]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    try:
        return len(workbook.sheetnames), list(workbook.sheetnames)
    finally:
        workbook.close()


def _save_uploaded_seed(root: Path, filename: str, content: bytes) -> Path:
    seed_dir = root / "data" / "raw" / "seed"
    seed_dir.mkdir(parents=True, exist_ok=True)
    safe_name = Path(filename).name
    if Path(safe_name).suffix.lower() != ".xlsx":
        raise ValueError("首次导入仅接受 .xlsx 文件。")
    digest = hashlib.sha256(content).hexdigest()
    target = seed_dir / safe_name
    if target.exists():
        if hashlib.sha256(target.read_bytes()).hexdigest() == digest:
            return target
        target = seed_dir / f"{target.stem}_{digest[:8]}{target.suffix}"
    if not target.exists():
        temporary = target.with_suffix(target.suffix + ".upload.tmp")
        temporary.write_bytes(content)
        os.replace(temporary, target)
    return target


def render_setup_wizard(root: Path) -> None:
    from policydb.ingest.excel import import_excel
    from policydb.query.database import build_database
    from policydb.settings import Settings

    status = initial_setup_status(root)
    st.title("首次设置向导")
    st.caption("数据库尚未准备完整。系统不会猜测或自动导入任何文件。")

    left, right = st.columns(2)
    left.metric("DuckDB", "已就绪" if status["database_ready"] else "缺失")
    right.metric("Curated数据", "已就绪" if status["curated_ready"] else "缺失")

    if status["curated_ready"] and not status["database_ready"]:
        st.info("已发现标准化数据，可以安全地从 Curated 层重建 DuckDB。")
        if st.button("从现有 Curated 数据重建数据库", type="primary"):
            try:
                with st.spinner("正在重建数据库……"):
                    build_database(Settings.discover(root))
                st.success("数据库重建完成，正在进入网站。")
                st.cache_resource.clear()
                st.rerun()
            except Exception as exc:
                st.error(f"数据库重建失败：{exc}")
        return

    st.subheader("选择种子 Excel")
    existing = sorted((root / "data" / "raw" / "seed").glob("*.xlsx"))
    source_mode = st.radio(
        "文件来源",
        ["选择Raw层已有文件", "从电脑上传文件"],
        horizontal=True,
        disabled=not existing,
    ) if existing else "从电脑上传文件"

    selected_path: Path | None = None
    uploaded_content: bytes | None = None
    display_name = ""
    if source_mode == "选择Raw层已有文件":
        selected_path = st.selectbox(
            "已有文件",
            existing,
            format_func=lambda path: path.name,
        )
        uploaded_content = selected_path.read_bytes()
        display_name = selected_path.name
    else:
        uploaded = st.file_uploader("上传原始政策数据库（.xlsx）", type=["xlsx"])
        if uploaded is not None:
            uploaded_content = uploaded.getvalue()
            display_name = Path(uploaded.name).name

    if uploaded_content is None:
        st.info("请选择明确的种子文件后继续。")
        return

    digest = hashlib.sha256(uploaded_content).hexdigest()
    try:
        sheet_count, sheet_names = _workbook_summary(uploaded_content)
    except Exception as exc:
        st.error(f"文件不是可读取的 Excel 工作簿：{exc}")
        return

    st.write(f"文件：`{display_name}`")
    st.write(f"SHA-256：`{digest}`")
    st.write(f"工作表数量：**{sheet_count}**")
    with st.expander("查看工作表名称"):
        st.write(sheet_names)

    confirmed = st.checkbox("我已确认这是需要导入的原始政策数据库文件")
    if st.button("确认文件并建立数据库", type="primary", disabled=not confirmed):
        try:
            with st.spinner("正在保存Raw副本、迁移工作表并建立数据库，请勿关闭页面……"):
                if selected_path is None:
                    selected_path = _save_uploaded_seed(root, display_name, uploaded_content)
                manifest = import_excel(selected_path, Settings.discover(root))
            st.success(
                f"导入完成：{manifest['sheet_count']}张工作表，"
                f"{manifest['record_count']}条记录。正在进入网站。"
            )
            st.cache_resource.clear()
            st.rerun()
        except Exception as exc:
            st.error(f"首次导入失败，原文件未被覆盖：{exc}")
